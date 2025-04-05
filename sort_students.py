import openpyxl
import streamlit as st
import io
import base64
from openpyxl.utils import get_column_letter
from copy import copy

def sort_students_alphabetically(wb):
    """Sort students alphabetically by name in each sheet"""
    for sheet_name in wb.sheetnames:
        if sheet_name == "All Student Data":
            continue
            
        ws = wb[sheet_name]
        st.write(f"Processing sheet: {sheet_name}")

        # Skip the "Student Details" header row
        current_row = 7  # Start after headers

        # Find student sections
        student_sections = []

        while current_row <= ws.max_row:
            cell_a = ws.cell(row=current_row, column=1).value
            
            # Check if this is a student name row
            if cell_a and "Class:" not in str(cell_a) and "Programme:" not in str(cell_a) and "Index No:" not in str(cell_a):
                student_name = str(cell_a).strip()
                start_row = current_row

                # Find end of student section (next empty row or next student)
                end_row = start_row
                while end_row < ws.max_row:
                    next_row = end_row + 1
                    next_cell_a = ws.cell(row=next_row, column=1).value
                    
                    # Stop if we hit an empty row
                    if not next_cell_a:
                        break
                        
                    end_row = next_row

                # Store section data with all formatting
                section_data = []
                for row in range(start_row, end_row + 1):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell_info = {
                            'value': cell.value,
                            'font': copy(cell.font),
                            'border': copy(cell.border),
                            'fill': copy(cell.fill),
                            'number_format': cell.number_format,
                            'protection': copy(cell.protection),
                            'alignment': copy(cell.alignment)
                        }
                        row_data.append(cell_info)
                    section_data.append(row_data)

                student_sections.append({
                    'name': student_name,
                    'start_row': start_row,
                    'end_row': end_row,
                    'data': section_data,
                    'height': end_row - start_row + 1
                })
                
                st.write(f"Found student: {student_name}")
                current_row = end_row + 2  # Skip the empty row
            else:
                current_row += 1

        if student_sections:
            # Sort sections by name
            sorted_sections = sorted(student_sections, key=lambda x: x['name'])
            st.write(f"Sorting {len(sorted_sections)} students...")

            # Write sorted sections back to sheet
            current_row = 7  # Start after headers
            for section in sorted_sections:
                # Write section data
                for row_idx, row_data in enumerate(section['data']):
                    for col_idx, cell_info in enumerate(row_data, 1):
                        target_cell = ws.cell(row=current_row + row_idx, column=col_idx)
                        target_cell.value = cell_info['value']
                        target_cell.font = cell_info['font']
                        target_cell.border = cell_info['border']
                        target_cell.fill = cell_info['fill']
                        target_cell.number_format = cell_info['number_format']
                        target_cell.protection = cell_info['protection']
                        target_cell.alignment = cell_info['alignment']

                current_row += section['height']
                # Add empty row between sections
                current_row += 1

            st.write(f"Successfully sorted students in sheet: {sheet_name}")
        else:
            st.write(f"No student sections found in sheet: {sheet_name}")

    return wb

def get_download_link(wb, filename):
    """Generate a download link for the Excel file"""
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)
    b64 = base64.b64encode(virtual_file.read()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'

def main():
    st.set_page_config(page_title="Student Sorting Tool", layout="wide")
    st.title("Student Sorting Tool")
    st.write("Sort students alphabetically while preserving all data and formatting")

    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx"])

    if uploaded_file is not None:
        st.success("File uploaded successfully!")
        if st.button("Sort Students Alphabetically"):
            with st.spinner("Sorting students..."):
                try:
                    file_bytes = uploaded_file.read()
                    virtual_file = io.BytesIO(file_bytes)
                    wb = openpyxl.load_workbook(virtual_file)
                    wb = sort_students_alphabetically(wb)
                    download_link = get_download_link(wb, "sorted_students.xlsx")
                    st.markdown(download_link, unsafe_allow_html=True)
                    st.success("Students sorted successfully!")
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")
                    st.exception(e)

if __name__ == "__main__":
    main()
