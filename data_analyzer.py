import pandas as pd
import numpy as np
import streamlit as st
import matplotlib.pyplot as plt
import seaborn as sns
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go

def load_data(uploaded_file):
    """Load data from uploaded Excel or CSV file"""
    if uploaded_file.name.endswith('.xlsx') or uploaded_file.name.endswith('.xls'):
        df = pd.read_excel(uploaded_file)
    elif uploaded_file.name.endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    else:
        st.error("Unsupported file format. Please upload an Excel or CSV file.")
        return None
    
    # Check if required columns exist
    required_columns = ["Student Code", "Programme", "Full Name", "Gender", 
                        "Date of Birth", "Basic Index No.", "Elective Subjects"]
    
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        st.error(f"Missing required columns: {', '.join(missing_columns)}")
        return None
    
    # Process the Elective Subjects column to extract individual subjects
    if "Elective Subjects" in df.columns:
        # Create a list of all unique subjects
        all_subjects = []
        for subjects in df["Elective Subjects"].dropna():
            if isinstance(subjects, str):
                subject_list = [s.strip() for s in subjects.split(',')]
                all_subjects.extend(subject_list)
        
        unique_subjects = sorted(list(set(all_subjects)))
        
        # Create indicator columns for each subject
        for subject in unique_subjects:
            df[f"Takes_{subject}"] = df["Elective Subjects"].apply(
                lambda x: 1 if isinstance(x, str) and subject in [s.strip() for s in x.split(',')] else 0
            )
    
    return df

def analyze_gender_distribution(df):
    """Analyze gender distribution overall and by programme with interactive features"""
    st.subheader("Gender Distribution")
    
    # Add filters
    st.write("Filter the data to see how it affects gender distribution:")
    
    # Programme filter
    all_programmes = ["All"] + sorted(df["Programme"].unique().tolist())
    selected_programme = st.selectbox("Select Programme", all_programmes)
    
    # Apply filters
    filtered_df = df.copy()
    if selected_programme != "All":
        filtered_df = filtered_df[filtered_df["Programme"] == selected_programme]
    
    # Overall gender distribution
    gender_counts = filtered_df["Gender"].value_counts()
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("Overall Gender Distribution")
        # Use Plotly for interactive pie chart
        fig = px.pie(
            values=gender_counts.values,
            names=gender_counts.index,
            title=f"Gender Distribution {'' if selected_programme == 'All' else f'for {selected_programme}'}",
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        fig.update_traces(textinfo='percent+label')
        st.plotly_chart(fig)
    
    with col2:
        st.write("Gender Distribution by Programme")
        if selected_programme == "All":
            gender_by_programme = pd.crosstab(df["Programme"], df["Gender"])
            
            # Use Plotly for interactive bar chart
            fig = px.bar(
                gender_by_programme,
                barmode='group',
                title="Gender Distribution by Programme",
                labels={"value": "Count", "index": "Programme"}
            )
            fig.update_layout(xaxis_tickangle=-45)
            st.plotly_chart(fig)
        else:
            st.info(f"Showing data for {selected_programme} only. Select 'All' to see comparison across programmes.")
    
    # Display the data in a table
    st.write("Gender Distribution by Programme (Table)")
    
    if selected_programme == "All":
        gender_by_programme = pd.crosstab(df["Programme"], df["Gender"])
        gender_by_programme_pct = gender_by_programme.div(gender_by_programme.sum(axis=1), axis=0) * 100
        gender_by_programme_pct = gender_by_programme_pct.round(1)
        
        # Combine counts and percentages
        result_df = pd.DataFrame()
        for col in gender_by_programme.columns:
            result_df[f"{col} (Count)"] = gender_by_programme[col]
            result_df[f"{col} (%)"] = gender_by_programme_pct[col].apply(lambda x: f"{x}%")
        
        st.dataframe(result_df)
    else:
        # Show simplified table for single programme
        gender_counts = filtered_df["Gender"].value_counts().reset_index()
        gender_counts.columns = ["Gender", "Count"]
        gender_counts["Percentage"] = (gender_counts["Count"] / gender_counts["Count"].sum() * 100).round(1).astype(str) + "%"
        st.dataframe(gender_counts)

def analyze_subjects(df):
    """Analyze elective subject distribution with interactive features"""
    st.subheader("Elective Subject Analysis")
    
    # Get all the subject indicator columns
    subject_columns = [col for col in df.columns if col.startswith("Takes_")]
    
    if not subject_columns:
        st.warning("No subject data to analyze.")
        return
    
    # Extract the subject names from the column names
    subjects = [col.replace("Takes_", "") for col in subject_columns]
    
    # Add filters
    st.write("Filter the data to customize your analysis:")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Programme filter
        all_programmes = ["All"] + sorted(df["Programme"].unique().tolist())
        selected_programme = st.selectbox("Select Programme", all_programmes, key="subject_programme")
    
    with col2:
        # Gender filter
        all_genders = ["All"] + sorted(df["Gender"].unique().tolist())
        selected_gender = st.selectbox("Select Gender", all_genders)
    
    # Number of subjects to display
    num_subjects = st.slider("Number of subjects to display", min_value=1, max_value=len(subjects), value=min(10, len(subjects)))
    
    # Sort options
    sort_options = ["Most Popular", "Least Popular", "Alphabetical"]
    sort_by = st.radio("Sort subjects by:", sort_options, horizontal=True)
    
    # Apply filters
    filtered_df = df.copy()
    if selected_programme != "All":
        filtered_df = filtered_df[filtered_df["Programme"] == selected_programme]
    if selected_gender != "All":
        filtered_df = filtered_df[filtered_df["Gender"] == selected_gender]
    
    # Count how many students take each subject
    subject_counts = filtered_df[subject_columns].sum()
    subject_counts.index = [idx.replace("Takes_", "") for idx in subject_counts.index]
    
    # Sort based on user selection
    if sort_by == "Most Popular":
        subject_counts = subject_counts.sort_values(ascending=False)
    elif sort_by == "Least Popular":
        subject_counts = subject_counts.sort_values(ascending=True)
    else:  # Alphabetical
        subject_counts = subject_counts.sort_index()
    
    # Limit to selected number of subjects
    subject_counts = subject_counts.head(num_subjects)
    
    # Plot the subject distribution using Plotly
    title_suffix = ""
    if selected_programme != "All":
        title_suffix += f" in {selected_programme}"
    if selected_gender != "All":
        title_suffix += f" for {selected_gender} Students"
    
    fig = px.bar(
        x=subject_counts.index,
        y=subject_counts.values,
        title=f"Number of Students Taking Each Elective Subject{title_suffix}",
        labels={"x": "Subject", "y": "Number of Students"},
        color=subject_counts.values,
        color_continuous_scale="Viridis"
    )
    fig.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig)
    
    # Display the data in a table
    st.write("Subject Distribution (Table)")
    subject_table = pd.DataFrame({
        "Subject": subject_counts.index,
        "Number of Students": subject_counts.values,
        "Percentage": (subject_counts.values / len(filtered_df) * 100).round(1).astype(str) + "%"
    })
    st.dataframe(subject_table)
    
    # Subject combinations analysis
    st.subheader("Subject Combinations Analysis")
    st.write("Analyze which subjects are commonly taken together")
    
    # Let user select a subject to analyze combinations
    selected_subject = st.selectbox("Select a subject to see what other subjects students commonly take with it:", 
                                   ["Select a subject..."] + subjects)
    
    if selected_subject != "Select a subject...":
        # Filter students who take the selected subject
        subject_takers = filtered_df[filtered_df[f"Takes_{selected_subject}"] == 1]
        
        if len(subject_takers) > 0:
            st.write(f"{len(subject_takers)} students take {selected_subject}")
            
            # Calculate what other subjects these students take
            other_subjects = [col for col in subject_columns if col != f"Takes_{selected_subject}"]
            other_subject_counts = subject_takers[other_subjects].sum()
            other_subject_counts.index = [idx.replace("Takes_", "") for idx in other_subject_counts.index]
            other_subject_counts = other_subject_counts.sort_values(ascending=False)
            
            # Calculate percentages
            other_subject_percentages = (other_subject_counts / len(subject_takers) * 100).round(1)
            
            # Create a horizontal bar chart
            fig = px.bar(
                y=other_subject_counts.index,
                x=other_subject_percentages.values,
                title=f"Subjects taken with {selected_subject} (% of {selected_subject} students)",
                labels={"y": "Subject", "x": "Percentage of Students"},
                orientation='h',
                color=other_subject_percentages.values,
                color_continuous_scale="Viridis"
            )
            fig.update_layout(yaxis={'categoryorder':'total ascending'})
            st.plotly_chart(fig)
            
            # Display the data in a table
            combo_table = pd.DataFrame({
                "Subject": other_subject_counts.index,
                "Count": other_subject_counts.values,
                "Percentage": other_subject_percentages.astype(str) + "%"
            })
            st.dataframe(combo_table)
            
            # Add student search and view functionality
            st.subheader("View Students Taking Specific Subject Combinations")
            
            # Let user select a second subject to find students taking both
            second_subject = st.selectbox(
                f"Select another subject to find students taking both {selected_subject} and this subject:",
                ["Any subject"] + [s for s in subjects if s != selected_subject]
            )
            
            # Filter students based on selection
            if second_subject == "Any subject":
                students_with_combo = subject_takers
                combo_description = f"taking {selected_subject}"
            else:
                students_with_combo = subject_takers[subject_takers[f"Takes_{second_subject}"] == 1]
                combo_description = f"taking both {selected_subject} and {second_subject}"
            
            # Display count of students with this combination
            st.write(f"Found {len(students_with_combo)} students {combo_description}")
            
            # Add search functionality
            search_term = st.text_input("Search students by name or student code:")
            
            if search_term:
                # Search in Full Name and Student Code columns
                search_results = students_with_combo[
                    students_with_combo["Full Name"].str.contains(search_term, case=False, na=False) |
                    students_with_combo["Student Code"].str.contains(search_term, case=False, na=False)
                ]
                
                if len(search_results) > 0:
                    st.write(f"Found {len(search_results)} matching students:")
                    st.dataframe(search_results[["Student Code", "Full Name", "Gender", "Programme"]])
                else:
                    st.info(f"No students found matching '{search_term}'")
            else:
                # Show all students with pagination
                students_per_page = st.slider("Students per page:", 5, 50, 10)
                total_pages = max(1, (len(students_with_combo) + students_per_page - 1) // students_per_page)
                
                if total_pages > 1:
                    page_number = st.number_input("Page:", min_value=1, max_value=total_pages, value=1)
                    start_idx = (page_number - 1) * students_per_page
                    end_idx = min(start_idx + students_per_page, len(students_with_combo))
                    
                    st.write(f"Showing students {start_idx+1}-{end_idx} of {len(students_with_combo)}")
                    st.dataframe(students_with_combo.iloc[start_idx:end_idx][["Student Code", "Full Name", "Gender", "Programme"]])
                    
                    st.write(f"Page {page_number} of {total_pages}")
                else:
                    st.dataframe(students_with_combo[["Student Code", "Full Name", "Gender", "Programme"]])
            
            # Option to view detailed student information
            st.subheader("View Detailed Student Information")
            st.write("Select a student to view all their information:")
            
            # Create a selectbox with student names and codes for easy identification
            student_options = ["Select a student..."] + [
                f"{row['Student Code']} - {row['Full Name']}" 
                for _, row in students_with_combo.iterrows()
            ]
            
            selected_student_option = st.selectbox("Select a student:", student_options)
            
            if selected_student_option != "Select a student...":
                # Extract student code from the selection
                selected_student_code = selected_student_option.split(" - ")[0].strip()
                
                # Convert both to strings and strip whitespace for comparison
                matching_students = students_with_combo[
                    students_with_combo["Student Code"].astype(str).str.strip() == str(selected_student_code)
                ]
                
                if len(matching_students) > 0:
                    student_data = matching_students.iloc[0]
                    
                    # Display student information in a formatted way
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.write("**Personal Information:**")
                        st.write(f"**Name:** {student_data['Full Name']}")
                        st.write(f"**Student Code:** {student_data['Student Code']}")
                        st.write(f"**Gender:** {student_data['Gender']}")
                        st.write(f"**Date of Birth:** {student_data['Date of Birth']}")
                        st.write(f"**Basic Index No.:** {student_data['Basic Index No.']}")
                    
                    with col2:
                        st.write("**Academic Information:**")
                        st.write(f"**Programme:** {student_data['Programme']}")
                        
                        # Get all subjects this student is taking
                        student_subjects = []
                        for col in subject_columns:
                            if student_data[col] == 1:
                                student_subjects.append(col.replace("Takes_", ""))
                        
                        st.write("**Elective Subjects:**")
                        for subject in student_subjects:
                            st.write(f"- {subject}")
                    
                    # Option to download this student's information
                    student_df = pd.DataFrame([student_data])
                    csv = student_df.to_csv(index=False)
                    b64 = base64.b64encode(csv.encode()).decode()
                    href = f'<a href="data:file/csv;base64,{b64}" download="{student_data["Student Code"]}_info.csv">Download Student Information</a>'
                    st.markdown(href, unsafe_allow_html=True)
                else:
                    st.error(f"No student found with code {selected_student_code}. This may be due to filtering applied to the data.")
        else:
            st.info(f"No students in the filtered dataset take {selected_subject}")

def analyze_age_distribution(df):
    """Analyze age distribution based on Date of Birth with interactive features"""
    st.subheader("Age Distribution")
    
    # Check if Date of Birth column exists and has valid dates
    if "Date of Birth" not in df.columns:
        st.warning("Date of Birth column not found.")
        return
    
    # Convert Date of Birth to datetime
    try:
        # Create a copy to avoid modifying the original dataframe
        age_df = df.copy()
        age_df["Date of Birth"] = pd.to_datetime(age_df["Date of Birth"], errors='coerce')
        
        # Calculate age
        today = pd.Timestamp.today()
        age_df["Age"] = (today - age_df["Date of Birth"]).dt.days // 365
        
        # Remove invalid ages
        age_df = age_df[age_df["Age"] > 0]
        
        if age_df["Age"].count() == 0:
            st.warning("No valid age data found.")
            return
        
        # Add filters
        st.write("Filter the data to customize your analysis:")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Programme filter
            all_programmes = ["All"] + sorted(age_df["Programme"].unique().tolist())
            selected_programme = st.selectbox("Select Programme", all_programmes, key="age_programme")
        
        with col2:
            # Gender filter
            all_genders = ["All"] + sorted(age_df["Gender"].unique().tolist())
            selected_gender = st.selectbox("Select Gender", all_genders, key="age_gender")
        
        # Apply filters
        filtered_df = age_df.copy()
        if selected_programme != "All":
            filtered_df = filtered_df[filtered_df["Programme"] == selected_programme]
        if selected_gender != "All":
            filtered_df = filtered_df[filtered_df["Gender"] == selected_gender]
        
        # Age statistics
        st.write(f"Average Age: {filtered_df['Age'].mean():.1f} years")
        st.write(f"Minimum Age: {filtered_df['Age'].min()} years")
        st.write(f"Maximum Age: {filtered_df['Age'].max()} years")
        
        # Age distribution visualization options
        viz_type = st.radio("Select visualization type:", ["Histogram", "Box Plot", "Violin Plot"], horizontal=True)
        
        if viz_type == "Histogram":
            # Interactive histogram with Plotly
            bin_size = st.slider("Bin Size (years)", min_value=1, max_value=5, value=1)
            
            fig = px.histogram(
                filtered_df, 
                x="Age",
                nbins=int((filtered_df["Age"].max() - filtered_df["Age"].min()) / bin_size),
                title="Age Distribution",
                labels={"Age": "Age (years)", "count": "Number of Students"},
                color_discrete_sequence=["#6495ED"]
            )
            fig.update_layout(bargap=0.1)
            st.plotly_chart(fig)
            
        elif viz_type == "Box Plot":
            # Choose grouping variable
            group_by = st.radio("Group by:", ["None", "Gender", "Programme"], horizontal=True)
            
            if group_by == "None":
                fig = px.box(
                    filtered_df,
                    y="Age",
                    title="Age Distribution",
                    labels={"Age": "Age (years)"}
                )
                st.plotly_chart(fig)
            else:
                fig = px.box(
                    filtered_df,
                    x=group_by,
                    y="Age",
                    title=f"Age Distribution by {group_by}",
                    labels={"Age": "Age (years)", group_by: group_by}
                )
                fig.update_layout(xaxis_tickangle=-45)
                st.plotly_chart(fig)
                
        else:  # Violin Plot
            # Choose grouping variable
            group_by = st.radio("Group by:", ["None", "Gender", "Programme"], horizontal=True, key="violin_group")
            
            if group_by == "None":
                fig = px.violin(
                    filtered_df,
                    y="Age",
                    title="Age Distribution",
                    labels={"Age": "Age (years)"},
                    box=True
                )
                st.plotly_chart(fig)
            else:
                fig = px.violin(
                    filtered_df,
                    x=group_by,
                    y="Age",
                    title=f"Age Distribution by {group_by}",
                    labels={"Age": "Age (years)", group_by: group_by},
                    box=True
                )
                fig.update_layout(xaxis_tickangle=-45)
                st.plotly_chart(fig)
        
        # Age distribution table
        st.write("Age Distribution Table")
        
        # Create age groups
        age_bins = [0, 15, 18, 21, 25, 30, 100]
        age_labels = ['Under 15', '15-17', '18-20', '21-24', '25-29', '30+']
        
        filtered_df['Age Group'] = pd.cut(filtered_df['Age'], bins=age_bins, labels=age_labels, right=False)
        
        # Count by age group
        age_group_counts = filtered_df['Age Group'].value_counts().sort_index()
        
        # Create a table
        age_table = pd.DataFrame({
            "Age Group": age_group_counts.index,
            "Count": age_group_counts.values,
            "Percentage": (age_group_counts.values / age_group_counts.sum() * 100).round(1).astype(str) + "%"
        })
        
        st.dataframe(age_table)
        
    except Exception as e:
        st.error(f"Error analyzing age distribution: {str(e)}")
        st.error("Please ensure the Date of Birth column contains valid dates.")

def generate_report(df):
    """Generate a comprehensive analysis report"""
    # Create a workbook
    wb = Workbook()
    
    # Create the Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    # Add title
    summary_ws['A1'] = "Student Data Analysis Report"
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws.merge_cells('A1:E1')
    summary_ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add basic statistics
    summary_ws['A3'] = "Total Number of Students:"
    summary_ws['B3'] = len(df)
    
    summary_ws['A4'] = "Number of Programmes:"
    summary_ws['B4'] = df["Programme"].nunique()
    
    summary_ws['A5'] = "Gender Distribution:"
    gender_counts = df["Gender"].value_counts()
    for i, (gender, count) in enumerate(gender_counts.items()):
        summary_ws[f'A{6+i}'] = f"    {gender}:"
        summary_ws[f'B{6+i}'] = count
        summary_ws[f'C{6+i}'] = f"{count/len(df)*100:.1f}%"
    
    # Add programme distribution
    programme_row = 8 + len(gender_counts)
    summary_ws[f'A{programme_row}'] = "Programme Distribution:"
    programme_counts = df["Programme"].value_counts()
    for i, (programme, count) in enumerate(programme_counts.items()):
        summary_ws[f'A{programme_row+1+i}'] = f"    {programme}:"
        summary_ws[f'B{programme_row+1+i}'] = count
        summary_ws[f'C{programme_row+1+i}'] = f"{count/len(df)*100:.1f}%"
    
    # Create Gender Analysis sheet
    gender_ws = wb.create_sheet(title="Gender Analysis")
    
    gender_ws['A1'] = "Gender Distribution by Programme"
    gender_ws['A1'].font = Font(bold=True, size=14)
    gender_ws.merge_cells('A1:E1')
    gender_ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add gender by programme data
    gender_by_programme = pd.crosstab(df["Programme"], df["Gender"])
    gender_by_programme_pct = gender_by_programme.div(gender_by_programme.sum(axis=1), axis=0) * 100
    
    # Write headers
    gender_ws['A3'] = "Programme"
    gender_ws['A3'].font = Font(bold=True)
    
    col_idx = 1
    for gender in gender_by_programme.columns:
        gender_ws.cell(row=3, column=col_idx+1).value = f"{gender} (Count)"
        gender_ws.cell(row=3, column=col_idx+1).font = Font(bold=True)
        gender_ws.cell(row=3, column=col_idx+2).value = f"{gender} (%)"
        gender_ws.cell(row=3, column=col_idx+2).font = Font(bold=True)
        col_idx += 2
    
    # Write data
    for i, programme in enumerate(gender_by_programme.index):
        gender_ws.cell(row=4+i, column=1).value = programme
        
        col_idx = 1
        for gender in gender_by_programme.columns:
            gender_ws.cell(row=4+i, column=col_idx+1).value = gender_by_programme.loc[programme, gender]
            gender_ws.cell(row=4+i, column=col_idx+2).value = f"{gender_by_programme_pct.loc[programme, gender]:.1f}%"
            col_idx += 2
    
    # Create Subject Analysis sheet
    subject_ws = wb.create_sheet(title="Subject Analysis")
    
    subject_ws['A1'] = "Elective Subject Distribution"
    subject_ws['A1'].font = Font(bold=True, size=14)
    subject_ws.merge_cells('A1:E1')
    subject_ws['A1'].alignment = Alignment(horizontal='center')
    
    # Get all the subject indicator columns
    subject_columns = [col for col in df.columns if col.startswith("Takes_")]
    
    if subject_columns:
        # Count how many students take each subject
        subject_counts = df[subject_columns].sum().sort_values(ascending=False)
        subject_counts.index = [idx.replace("Takes_", "") for idx in subject_counts.index]
        
        # Write headers
        subject_ws['A3'] = "Subject"
        subject_ws['A3'].font = Font(bold=True)
        subject_ws['B3'] = "Number of Students"
        subject_ws['B3'].font = Font(bold=True)
        subject_ws['C3'] = "Percentage"
        subject_ws['C3'].font = Font(bold=True)
        
        # Write data
        for i, (subject, count) in enumerate(subject_counts.items()):
            subject_ws.cell(row=4+i, column=1).value = subject
            subject_ws.cell(row=4+i, column=2).value = count
            subject_ws.cell(row=4+i, column=3).value = f"{count/len(df)*100:.1f}%"
    
    # Adjust column widths
    for sheet in wb.worksheets:
        for col in range(1, 10):  # Adjust up to column J
            max_length = 0
            column_cells = sheet[get_column_letter(col)]
            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 10)
            sheet.column_dimensions[get_column_letter(col)].width = adjusted_width
    
    # Save to a virtual file
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)
    
    return virtual_file

def get_download_link(virtual_file, filename):
    """Generate a download link for the Excel file"""
    b64 = base64.b64encode(virtual_file.read()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Analysis Report</a>'

def run_app():
    """Main application function without page config"""
    st.title("Student Data Analyzer")
    st.write("Upload your student data file for analysis")
    
    uploaded_file = st.file_uploader("Choose an Excel or CSV file", type=["xlsx", "xls", "csv"])
    
    if uploaded_file is not None:
        st.success("File uploaded successfully!")
        
        with st.spinner("Loading and processing data..."):
            df = load_data(uploaded_file)
            
            if df is not None:
                st.success(f"Data loaded successfully! Found {len(df)} student records.")
                
                # Display data preview with option to see more
                st.subheader("Data Preview")
                show_all_data = st.checkbox("Show all data")
                if show_all_data:
                    st.dataframe(df)
                else:
                    st.dataframe(df.head())
                
                # Data filtering options
                st.subheader("Data Filtering")
                st.write("Filter the data for all analyses below:")
                
                # Programme filter for all analyses
                all_programmes = ["All"] + sorted(df["Programme"].unique().tolist())
                global_programme = st.selectbox("Filter by Programme (applies to all analyses)", all_programmes, key="global_programme")
                
                # Gender filter for all analyses
                all_genders = ["All"] + sorted(df["Gender"].unique().tolist())
                global_gender = st.selectbox("Filter by Gender (applies to all analyses)", all_genders, key="global_gender")
                
                # Apply global filters
                filtered_df = df.copy()
                if global_programme != "All":
                    filtered_df = filtered_df[filtered_df["Programme"] == global_programme]
                if global_gender != "All":
                    filtered_df = filtered_df[filtered_df["Gender"] == global_gender]
                
                # Show filter summary
                filter_text = "Currently showing data for "
                if global_programme == "All" and global_gender == "All":
                    filter_text += "all students"
                else:
                    if global_gender != "All":
                        filter_text += f"{global_gender} students"
                    else:
                        filter_text += "all genders"
                    
                    if global_programme != "All":
                        filter_text += f" in {global_programme} programme"
                    else:
                        filter_text += " across all programmes"
                
                st.info(filter_text)
                
                # Basic statistics
                st.subheader("Basic Statistics")
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Students", len(filtered_df))
                with col2:
                    st.metric("Number of Programmes", filtered_df["Programme"].nunique())
                with col3:
                    male_count = len(filtered_df[filtered_df["Gender"] == "Male"]) if "Male" in filtered_df["Gender"].values else 0
                    female_count = len(filtered_df[filtered_df["Gender"] == "Female"]) if "Female" in filtered_df["Gender"].values else 0
                    
                    if male_count > 0 and female_count > 0:
                        ratio = female_count / male_count
                        st.metric("Gender Ratio (M:F)", f"1:{ratio:.2f}")
                    else:
                        st.metric("Gender Ratio (M:F)", "N/A")
                
                # Create tabs for different analyses
                tab1, tab2, tab3, tab4 = st.tabs(["Gender Analysis", "Subject Analysis", "Age Analysis", "Generate Report"])
                
                with tab1:
                    analyze_gender_distribution(filtered_df)
                
                with tab2:
                    analyze_subjects(filtered_df)
                
                with tab3:
                    analyze_age_distribution(filtered_df)
                
                with tab4:
                    st.subheader("Generate Comprehensive Analysis Report")
                    st.write("Click the button below to generate a detailed Excel report with all analyses.")
                    
                    # Options for the report
                    include_filtered_only = st.checkbox("Include only filtered data in report", value=True)
                    
                    if st.button("Generate Report"):
                        with st.spinner("Generating report..."):
                            report_df = filtered_df if include_filtered_only else df
                            report_file = generate_report(report_df)
                            download_link = get_download_link(report_file, "student_data_analysis.xlsx")
                            st.markdown(download_link, unsafe_allow_html=True)
                            st.success("Report generated successfully!")
                            
                            if include_filtered_only and (global_programme != "All" or global_gender != "All"):
                                st.info(f"Report includes only the filtered data ({len(filtered_df)} students).")
                            else:
                                st.info(f"Report includes all data ({len(df)} students).")

def main():
    """Entry point when running this file directly"""
    # Only set page config when running this file directly
    st.set_page_config(page_title="Student Data Analyzer", layout="wide")
    run_app()

if __name__ == "__main__":
    main()

