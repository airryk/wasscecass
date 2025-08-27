import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import io
import base64
from fpdf import FPDF
from streamlit_sortables import sort_items
import datetime

# Try to import plotly, handle if not available
try:
    import plotly.express as px
    PLOTLY_AVAILABLE = True
except ImportError:
    PLOTLY_AVAILABLE = False

class PDFWithPageNumber(FPDF):
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', 0, 0, 'C')

def create_class_list_pdf(arrangement_df, exam_date):
    """Generates a PDF class list for signing for a specific date."""
    pdf = PDFWithPageNumber(orientation='P', unit='mm', format='A4')
    
    date_df = arrangement_df[arrangement_df['Date'] == exam_date]
    if date_df.empty:
        return None
    
    for session in sorted(date_df['Session'].unique()):
        session_df = date_df[date_df['Session'] == session]
        rooms = sorted(session_df['Room'].unique())
        
        for room in rooms:
            pdf.add_page()
            
            pdf.set_font('Arial', 'B', 16)
            title_text = f"Exam List for {room} - {session} Session"
            pdf.cell(0, 10, title_text, 0, 1, 'C')
            
            exam_date_text = f"Exam Date: {exam_date.strftime('%A, %B %d, %Y')}"
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 10, exam_date_text, 0, 1, 'C')
            pdf.ln(5)
            
            pdf.set_font('Arial', 'B', 10)
            col_widths = [15, 25, 60, 25, 40, 30]
            headers = ['Seat No', 'Index Number', 'Full Name', 'Class', 'Subject', 'Signature']
            for i, header in enumerate(headers):
                pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
            pdf.ln()
            
            pdf.set_font('Arial', '', 8)
            room_df = session_df[session_df['Room'] == room].sort_values('Seat No')
            for _, row in room_df.iterrows():
                pdf.cell(col_widths[0], 10, str(row['Seat No']), 1)
                pdf.cell(col_widths[1], 10, str(row['Index Number']), 1)
                pdf.cell(col_widths[2], 10, str(row['Full Name']), 1)
                pdf.cell(col_widths[3], 10, str(row['Class']), 1)
                pdf.cell(col_widths[4], 10, str(row['Subject']), 1)
                pdf.cell(col_widths[5], 10, '', 1)
                pdf.ln()
                
    # Fix: Return bytes instead of string
    pdf_output = pdf.output(dest='S')
    return pdf_output.encode('latin-1') if isinstance(pdf_output, str) else pdf_output

def create_pdf(arrangement_df, exam_date):
    """Generates a PDF file from the arrangement dataframe for a specific date."""
    pdf = PDFWithPageNumber(orientation='L', unit='mm', format='A4')
    
    date_df = arrangement_df[arrangement_df['Date'] == exam_date]
    if date_df.empty:
        return None
    
    for session in sorted(date_df['Session'].unique()):
        pdf.add_page()
        
        pdf.set_font('Arial', 'B', 16)
        title_text = f"Seating Arrangement for {exam_date.strftime('%A, %B %d, %Y')} - {session} Session"
        pdf.cell(0, 10, title_text, 0, 1, 'C')
        
        pdf.set_font('Arial', 'B', 10)
        col_widths = [40, 20, 35, 65, 35, 40, 30]
        headers = ['Room', 'Seat No', 'Index Number', 'Full Name', 'Class', 'Subject', 'Session']
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
        pdf.ln()
        
        pdf.set_font('Arial', '', 8)
        session_df = date_df[date_df['Session'] == session]
        for _, row in session_df.iterrows():
            pdf.cell(col_widths[0], 10, str(row['Room']), 1)
            pdf.cell(col_widths[1], 10, str(row['Seat No']), 1)
            pdf.cell(col_widths[2], 10, str(row['Index Number']), 1)
            pdf.cell(col_widths[3], 10, str(row['Full Name']), 1)
            pdf.cell(col_widths[4], 10, str(row['Class']), 1)
            pdf.cell(col_widths[5], 10, str(row['Subject']), 1)
            pdf.cell(col_widths[6], 10, str(row['Session']), 1)
            pdf.ln()
            
    # Fix: Return bytes instead of string
    pdf_output = pdf.output(dest='S')
    return pdf_output.encode('latin-1') if isinstance(pdf_output, str) else pdf_output

def analyze_subject_registration(df):
    """Analyze subject registration by gender and return statistics."""
    if df is None or df.empty:
        return None
    
    # Check if Gender column exists
    if 'Gender' not in df.columns:
        st.warning("Gender column not found in the data. Gender analysis will not be available.")
        has_gender = False
    else:
        has_gender = True
    
    # Get all subjects from both core and elective columns
    all_subjects = set()
    
    if 'Core_Subjects' in df.columns:
        core_subjects = df['Core_Subjects'].str.split(',').explode().str.strip()
        all_subjects.update(core_subjects.dropna().unique())
    
    if 'Elective_Subjects' in df.columns:
        elective_subjects = df['Elective_Subjects'].str.split(',').explode().str.strip()
        all_subjects.update(elective_subjects.dropna().unique())
    
    # Remove empty strings and NaN values
    all_subjects = {s for s in all_subjects if s and s != 'nan' and str(s).strip()}
    
    if not all_subjects:
        return None
    
    # Create analysis data
    analysis_data = []
    
    for subject in sorted(all_subjects):
        # Find students taking this subject (exact match)
        core_mask = df['Core_Subjects'].str.split(',').apply(lambda x: subject in [s.strip() for s in x] if isinstance(x, list) else False)
        elective_mask = df['Elective_Subjects'].str.split(',').apply(lambda x: subject in [s.strip() for s in x] if isinstance(x, list) else False)
        subject_students = df[core_mask | elective_mask]
        
        total_count = len(subject_students)
        
        if has_gender and total_count > 0:
            # Count by gender
            gender_counts = subject_students['Gender'].value_counts()
            male_count = gender_counts.get('Male', 0) + gender_counts.get('M', 0) + gender_counts.get('male', 0)
            female_count = gender_counts.get('Female', 0) + gender_counts.get('F', 0) + gender_counts.get('female', 0)
            
            # Make total match the sum of male and female counts
            total_with_gender = male_count + female_count
            
            analysis_data.append({
                'Subject': subject,
                'Male': male_count,
                'Female': female_count,
                'Total': total_with_gender
            })
        else:
            analysis_data.append({
                'Subject': subject,
                'Total': total_count
            })
    
    return pd.DataFrame(analysis_data) if analysis_data else None

def display_subject_analysis(df):
    """Display subject analysis in the Streamlit app."""
    st.subheader("📊 Subject Registration Analysis")
    
    analysis_df = analyze_subject_registration(df)
    
    if analysis_df is None or analysis_df.empty:
        st.warning("No subject data available for analysis.")
        return
    
    # Display the analysis table
    if 'Male' in analysis_df.columns and 'Female' in analysis_df.columns:
        st.dataframe(
            analysis_df.style.format({
                'Male': '{:,}',
                'Female': '{:,}',
                'Total': '{:,}'
            }),
            use_container_width=True
        )
        
        # Create summary statistics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total Subjects", len(analysis_df))
        
        with col2:
            total_registrations = analysis_df['Total'].sum()
            st.metric("Total Registrations", f"{total_registrations:,}")
        
        with col3:
            total_male = analysis_df['Male'].sum()
            st.metric("Total Male", f"{total_male:,}")
        
        with col4:
            total_female = analysis_df['Female'].sum()
            st.metric("Total Female", f"{total_female:,}")
        
        # Show gender distribution chart
        if PLOTLY_AVAILABLE:
            if st.checkbox("Show Gender Distribution Chart"):
                # Create a chart showing gender distribution by subject
                chart_data = analysis_df.melt(
                    id_vars=['Subject'], 
                    value_vars=['Male', 'Female'], 
                    var_name='Gender', 
                    value_name='Count'
                )
                
                fig = px.bar(
                    chart_data, 
                    x='Subject', 
                    y='Count', 
                    color='Gender',
                    title='Student Registration by Subject and Gender',
                    color_discrete_map={'Male': '#1f77b4', 'Female': '#ff7f0e'}
                )
                fig.update_layout(xaxis_tickangle=45)
                st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("📊 Install plotly (`pip install plotly`) to enable interactive charts.")
    
    else:
        # Display without gender breakdown
        st.dataframe(
            analysis_df.style.format({'Total': '{:,}'}),
            use_container_width=True
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Total Subjects", len(analysis_df))
        
        with col2:
            total_registrations = analysis_df['Total'].sum()
            st.metric("Total Registrations", f"{total_registrations:,}")
    
    # Show top subjects
    if len(analysis_df) > 0:
        st.subheader("📈 Top 10 Most Popular Subjects")
        top_subjects = analysis_df.nlargest(10, 'Total')[['Subject', 'Total']]
        st.dataframe(top_subjects, use_container_width=True, hide_index=True)

def generate_arrangement(df, room_capacities, subject_details):
    """Generates an Excel workbook with seating arrangements for multiple subjects in rooms."""
    required_columns = ['IndexNumber', 'Full_Name', 'Core_Subjects', 'Elective_Subjects', 'Class']
    if not all(col in df.columns for col in required_columns):
        missing_cols = [col for col in required_columns if col not in df.columns]
        st.error(f"The uploaded file is missing required columns: {', '.join(missing_cols)}")
        return None, None, None

    student_subjects = []
    for _, row in df.iterrows():
        core_subjects = str(row.get('Core_Subjects', '')).split(',')
        elective_subjects = str(row.get('Elective_Subjects', '')).split(',')
        all_subjects_list = [s.strip() for s in core_subjects + elective_subjects if s.strip()]
        for subject in all_subjects_list:
            student_subjects.append({
                'IndexNumber': row['IndexNumber'],
                'Full_Name': row['Full_Name'],
                'Class': row['Class'],
                'Subject': subject
            })

    if not student_subjects:
        st.warning("No subjects could be processed from the uploaded file.")
        return None, None, None

    long_df = pd.DataFrame(student_subjects)
    
    selected_subjects = list(subject_details.keys())
    long_df = long_df[long_df['Subject'].isin(selected_subjects)]

    if long_df.empty:
        st.warning("No students found for the selected subjects.")
        return None, None, None

    long_df['Date'] = long_df['Subject'].map(lambda s: subject_details[s]['date'])
    long_df['Session'] = long_df['Subject'].map(lambda s: subject_details[s]['session'])

    both_session_df = long_df[long_df['Session'] == 'Both'].copy()
    if not both_session_df.empty:
        morning_df = both_session_df.copy()
        morning_df['Session'] = 'Morning'
        afternoon_df = both_session_df.copy()
        afternoon_df['Session'] = 'Afternoon'
        long_df = pd.concat([long_df[long_df['Session'] != 'Both'], morning_df, afternoon_df]).reset_index(drop=True)

    arrangement_df = pd.DataFrame()
    
    unique_dates = sorted(long_df['Date'].unique())

    for exam_date in unique_dates:
        date_df = long_df[long_df['Date'] == exam_date]

        for session in ['Morning', 'Afternoon']:
            session_df = date_df[date_df['Session'] == session]
            
            if session_df.empty:
                continue

            all_seats = []
            room_names = {room: f"Room {i+1} ({room})" for i, room in enumerate(room_capacities.keys())}
            for room_name_key in room_capacities.keys():
                capacity = room_capacities[room_name_key]
                for seat_num in range(1, capacity + 1):
                    all_seats.append({'Room': room_names[room_name_key], 'Seat No': seat_num})

            if len(session_df) > len(all_seats):
                st.error(f"Not enough seats for the {session} session on {exam_date.strftime('%Y-%m-%d')} ({len(session_df)} required, {len(all_seats)} available).")
                continue

            seating_arrangement = []
            seat_index = 0
            
            subject_counts = session_df['Subject'].value_counts()
            subjects_in_session = subject_counts.index.tolist()
            
            for i, subject in enumerate(subjects_in_session):
                subject_df = session_df[session_df['Subject'] == subject].sort_values('IndexNumber')
                for _, student in subject_df.iterrows():
                    if seat_index < len(all_seats):
                        seat = all_seats[seat_index]
                        seating_arrangement.append({
                            'Date': student['Date'],
                            'Room': seat['Room'],
                            'Seat No': seat['Seat No'],
                            'Index Number': student['IndexNumber'],
                            'Full Name': student['Full Name'],
                            'Class': student['Class'],
                            'Subject': student['Subject'],
                            'Session': student['Session']
                        })
                        seat_index += 1
                    else:
                        st.error(f"Ran out of seats during arrangement for the {session} session on {exam_date.strftime('%Y-%m-%d')}.")
                        break
                
                if seat_index >= len(all_seats) and i < len(subjects_in_session) - 1:
                    st.error(f"Not enough seats for all subjects in the {session} session on {exam_date.strftime('%Y-%m-%d')}.")
                    break
            
            if seating_arrangement:
                arrangement_df = pd.concat([arrangement_df, pd.DataFrame(seating_arrangement)])

    if arrangement_df.empty:
        st.warning("No seating arrangement could be generated.")
        return None, None, None
        
    arrangement_df['Date'] = pd.to_datetime(arrangement_df['Date']).dt.date

    # Calculate statistics
    stats = {}
    for exam_date in unique_dates:
        date_df = arrangement_df[arrangement_df['Date'] == exam_date]
        for session in ['Morning', 'Afternoon']:
            session_df = date_df[date_df['Session'] == session]
            if not session_df.empty:
                total_students = len(session_df)
                total_seats = sum(room_capacities.values())
                rooms_used = session_df['Room'].nunique()
                stats[f"{exam_date.strftime('%Y-%m-%d')} {session}"] = {
                    "total_students": total_students,
                    "total_seats": total_seats,
                    "rooms_used": rooms_used
                }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for exam_date in sorted(arrangement_df['Date'].unique()):
        ws = wb.create_sheet(title=exam_date.strftime('%Y-%m-%d'))
        
        title_text = f"Seating Arrangement for {exam_date.strftime('%A, %B %d, %Y')}"
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')

        headers = ['Room', 'Seat No', 'Index Number', 'Full Name', 'Class', 'Subject', 'Session']
        ws.append(headers)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        date_df = arrangement_df[arrangement_df['Date'] == exam_date]
        for r_idx, row in enumerate(date_df.itertuples(index=False), 3):
            for c_idx, value in enumerate(row[1:], 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 20

    return wb, arrangement_df, stats

def get_excel_download_link(wb, filename):
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)
    b64 = base64.b64encode(virtual_file.read()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'

def get_pdf_download_link(pdf_bytes, filename):
    b64 = base64.b64encode(pdf_bytes).decode()
    return f'<a href="data:application/pdf;base64,{b64}" download="{filename}">Download PDF file</a>'

def sync_ordered_rooms():
    if 'selected_rooms' not in st.session_state:
        return

    if 'ordered_rooms' not in st.session_state:
        st.session_state.ordered_rooms = st.session_state.selected_rooms.copy()
        return

    ordered = st.session_state.ordered_rooms
    selected = st.session_state.selected_rooms

    new_ordered = [room for room in ordered if room in selected]

    for room in selected:
        if room not in new_ordered:
            new_ordered.append(room)
    
    st.session_state.ordered_rooms = new_ordered

def add_room_callback():
    new_room = st.session_state.get("new_room_input", "").strip()
    if new_room and new_room not in st.session_state.all_classes:
        st.session_state.all_classes.append(new_room)
        st.session_state.all_classes.sort()
        if 'selected_rooms' in st.session_state:
            st.session_state.selected_rooms.append(new_room)
        st.session_state.new_room_input = ""
        sync_ordered_rooms()

def create_template_file():
    """Creates a template Excel file with the required columns and example data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Define headers
    headers = ['IndexNumber', 'Full_Name', 'Class', 'Gender', 'Core_Subjects', 'Elective_Subjects']
    ws.append(headers)
    
    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add example data
    example_data = [
        ['001', 'John Doe', '3A', 'Male', 'Mathematics,English,Science', 'History,Geography'],
        ['002', 'Jane Smith', '3A', 'Female', 'Mathematics,English,Science', 'French,Art'],
        ['003', 'Alex Johnson', '3B', 'Male', 'Mathematics,English,Science', 'Computer Science,Physics'],
        ['004', 'Sarah Williams', '3B', 'Female', 'Mathematics,English,Science', 'Chemistry,Biology'],
        ['005', 'Michael Brown', '3C', 'Male', 'Mathematics,English,Science', 'Economics,Business Studies']
    ]
    
    for row in example_data:
        ws.append(row)
    
    # Add instructions in a new sheet
    ws_instructions = wb.create_sheet(title="Instructions")
    instructions = [
        ["Instructions for filling the template:"],
        [""],
        ["1. IndexNumber: Unique identifier for each student (required)"],
        ["2. Full_Name: Complete name of the student (required)"],
        ["3. Class: The class or group the student belongs to (required)"],
        ["4. Gender: Student's gender (optional, but useful for analysis)"],
        ["5. Core_Subjects: List of core subjects separated by commas (required)"],
        ["6. Elective_Subjects: List of elective subjects separated by commas (required)"],
        [""],
        ["Notes:"],
        ["- Do not change the column headers"],
        ["- Make sure subject names are consistent (e.g., 'Mathematics' vs 'Math')"],
        ["- You can add as many rows as needed"],
        ["- Save the file as Excel (.xlsx) or CSV format before uploading"]
    ]
    
    for row in instructions:
        ws_instructions.append(row)
    
    # Format the instructions
    ws_instructions.column_dimensions['A'].width = 80
    for i, row in enumerate(ws_instructions.iter_rows(min_row=1, max_row=len(instructions)), 1):
        if i == 1:  # Title row
            row[0].font = Font(bold=True, size=14)
        elif i > 2 and i <= 8:  # Column descriptions
            row[0].font = Font(bold=False)
        elif i > 9:  # Notes
            row[0].font = Font(italic=True)
    
    # Set column widths in template sheet
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    
    return wb

def run_app():
    st.title("Exam Seating Arrangement Generator")
    st.write("Upload a file with student data to generate a seating arrangement for exams based on subjects.")
    
    # Add template download section
    st.subheader("Need a template?")
    if st.button("Generate Template File"):
        template_wb = create_template_file()
        template_file = io.BytesIO()
        template_wb.save(template_file)
        template_file.seek(0)
        
        b64_template = base64.b64encode(template_file.read()).decode()
        template_download_link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_template}" download="seating_arrangement_template.xlsx">Download Template Excel File</a>'
        st.markdown(template_download_link, unsafe_allow_html=True)
        st.info("Template file generated! Click the link above to download.")
        st.markdown("---")
    
    uploaded_file = st.file_uploader(
        "Upload student data file (CSV or Excel)",
        type=["csv", "xlsx"],
        help="The file must contain columns: IndexNumber, Full_Name, Class, Core_Subjects, Elective_Subjects. Optional: Gender column for gender analysis."
    )

    if uploaded_file:
        try:
            df = pd.read_excel(uploaded_file, dtype=str) if uploaded_file.name.endswith('.xlsx') else pd.read_csv(uploaded_file, dtype=str)
            st.success("File uploaded successfully!")
            
            # Create tabs for different views
            tab1, tab2 = st.tabs(["📋 Data Preview", "📊 Subject Analysis"])
            
            with tab1:
                st.dataframe(df.head())
            
            with tab2:
                display_subject_analysis(df)

            st.header("Exam Configuration")

            if 'Core_Subjects' in df.columns and 'Elective_Subjects' in df.columns:
                core_subjects = df['Core_Subjects'].str.split(',').explode().str.strip().astype(str).unique()
                elective_subjects = df['Elective_Subjects'].str.split(',').explode().str.strip().astype(str).unique()
                all_subjects = sorted(list(set(core_subjects) | set(elective_subjects)))
                
                # Remove empty strings and NaN values
                all_subjects = [s for s in all_subjects if s and s != 'nan' and str(s).strip()]
                
                selected_subjects = st.multiselect("Select subjects for this exam session", options=all_subjects)

                subject_details = {}
                if selected_subjects:
                    st.subheader("Set Subject Dates and Sessions")
                    today = datetime.date.today()
                    
                    # Show selected subjects analysis
                    if st.checkbox("Show analysis for selected subjects only"):
                        selected_analysis = analyze_subject_registration(df)
                        if selected_analysis is not None:
                            selected_analysis = selected_analysis[selected_analysis['Subject'].isin(selected_subjects)]
                            if not selected_analysis.empty:
                                st.subheader("📊 Selected Subjects Analysis")
                                if 'Male' in selected_analysis.columns and 'Female' in selected_analysis.columns:
                                    st.dataframe(
                                        selected_analysis.style.format({
                                            'Male': '{:,}',
                                            'Female': '{:,}',
                                            'Total': '{:,}'
                                        }),
                                        use_container_width=True
                                    )
                                    
                                    col1, col2, col3 = st.columns(3)
                                    with col1:
                                        st.metric("Selected Subjects", len(selected_analysis))
                                    with col2:
                                        st.metric("Total Male Students", f"{selected_analysis['Male'].sum():,}")
                                    with col3:
                                        st.metric("Total Female Students", f"{selected_analysis['Female'].sum():,}")
                                else:
                                    st.dataframe(
                                        selected_analysis.style.format({'Total': '{:,}'}),
                                        use_container_width=True
                                    )
                                    col1, col2 = st.columns(2)
                                    with col1:
                                        st.metric("Selected Subjects", len(selected_analysis))
                                    with col2:
                                        st.metric("Total Students", f"{selected_analysis['Total'].sum():,}")
                    
                    for subject in selected_subjects:
                        st.write(f"**{subject}**")
                        cols = st.columns([0.5, 0.5])
                        date = cols[0].date_input("Date", value=today, key=f"date_{subject}")
                        session = cols[1].selectbox(f"Session for {subject}", ["Morning", "Afternoon", "Both"], key=f"session_{subject}")
                        subject_details[subject] = {'date': date, 'session': session}
            else:
                st.error("The file must contain 'Core_Subjects' and 'Elective_Subjects' columns.")
                return

            if 'Class' in df.columns:
                if 'current_file' not in st.session_state or st.session_state.current_file != uploaded_file.name:
                    st.session_state.current_file = uploaded_file.name
                    all_classes = sorted(df['Class'].unique())
                    st.session_state.all_classes = all_classes
                    st.session_state.ordered_rooms = all_classes.copy()
                    st.session_state.selected_rooms = all_classes.copy()

                st.text_input("Add a new room (optional)", key="new_room_input")
                st.button("Add Room", on_click=add_room_callback)

                st.subheader("Manage and Order Rooms")
                st.multiselect(
                    "Select classes to use as rooms",
                    options=st.session_state.all_classes,
                    key="selected_rooms",
                    on_change=sync_ordered_rooms
                )

                st.write("Order of Rooms (Drag to reorder, top has higher priority):")
                if 'ordered_rooms' in st.session_state and st.session_state.ordered_rooms:
                    sorted_rooms = sort_items(st.session_state.ordered_rooms, direction='vertical')
                    if sorted_rooms != st.session_state.ordered_rooms:
                        st.session_state.ordered_rooms = sorted_rooms
                        st.rerun()
            else:
                st.error("The uploaded file must contain a 'Class' column.")
                return

            room_capacities = {}
            if st.session_state.get('ordered_rooms'):
                st.subheader("Set Room Capacities")
                for room in st.session_state.ordered_rooms:
                    room_capacities[room] = st.number_input(f"Capacity for {room}", min_value=1, value=30, key=f"capacity_{room}")

            if st.button("Generate Seating Arrangement"):
                if not st.session_state.get('ordered_rooms'):
                    st.warning("Please select at least one class to be used as a room.")
                elif not selected_subjects:
                    st.warning("Please select at least one subject for the session.")
                else:
                    with st.spinner("Generating arrangement..."):
                        wb, arrangement_df, stats = generate_arrangement(df, room_capacities, subject_details)
                        if wb:
                            st.success("Seating arrangement generated successfully!")

                            st.subheader("Arrangement Statistics")
                            if stats:
                                for session, data in stats.items():
                                    st.write(f"**{session}:**")
                                    st.write(f"  - Total Students: {data['total_students']}")
                                    st.write(f"  - Total Seats Available: {data['total_seats']}")
                                    st.write(f"  - Rooms Used: {data['rooms_used']}")
                            
                            subjects_in_arrangement = arrangement_df['Subject'].unique()
                            subjects_str = "_".join(subjects_in_arrangement).replace(" ", "_")
                            excel_filename = f"seating_arrangement_{subjects_str}.xlsx"
                            st.markdown(get_excel_download_link(wb, excel_filename), unsafe_allow_html=True)

                            st.subheader("Download PDFs by Date")
                            if arrangement_df is not None and not arrangement_df.empty:
                                for exam_date in sorted(arrangement_df['Date'].unique()):
                                    st.write(f"**Downloads for {exam_date.strftime('%A, %B %d, %Y')}:**")

                                    subjects_on_date = arrangement_df[arrangement_df['Date'] == exam_date]['Subject'].unique()
                                    subjects_on_date_str = '_'.join(subjects_on_date).replace(' ', '_')
                                    
                                    pdf_bytes = create_pdf(arrangement_df, exam_date)
                                    if pdf_bytes:
                                        pdf_filename = f"seating_arrangement_{subjects_on_date_str}_{exam_date.strftime('%Y-%m-%d')}.pdf"
                                        st.markdown(get_pdf_download_link(pdf_bytes, pdf_filename), unsafe_allow_html=True)

                                    class_list_pdf_bytes = create_class_list_pdf(arrangement_df, exam_date)
                                    if class_list_pdf_bytes:
                                        class_list_pdf_filename = f"exam_list_{subjects_on_date_str}_{exam_date.strftime('%Y-%m-%d')}.pdf"
                                        st.markdown(get_pdf_download_link(class_list_pdf_bytes, class_list_pdf_filename).replace("Download PDF file", "Download Exam List PDF"), unsafe_allow_html=True)

        except Exception as e:
            st.error(f"An error occurred while processing the file: {e}")

    # Add footer with link to services
    st.markdown("---")
    st.markdown(
        """
        <div style="text-align: center; padding: 10px;">
            <p>For more educational tools and services, visit <a href="https://techmawu.com" target="_blank">techmawu.com</a></p>
        </div>
        """, 
        unsafe_allow_html=True
    )

def main():
    """Entry point when running this file directly"""
    st.set_page_config(page_title="Exam Seating Arrangement Generator", layout="wide")
    run_app()

if __name__ == "__main__":
    main()
