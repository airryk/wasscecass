import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
import random
import streamlit as st
import io
import base64
from openpyxl.utils import get_column_letter

def generate_student_data(num_students=20):
    """Generate random student data"""
    classes = ["3A1", "3A2", "3S1", "3S2", "3B1"]
    programmes = ["GENERAL ARTS", "SCIENCE", "BUSINESS", "HOME ECONOMICS"]
    
    # Define possible subjects for each category
    c_subjects = ["C-MATHS", "ENG", "SOC-STUD", "INT-SCI"]
    e_subjects = ["GOV", "ECONS", "LIT-ENG", "FANTE", "PHYSICS", "CHEMISTRY", "BIOLOGY", "GEOGRAPHY"]
    
    data = []
    for i in range(1, num_students + 1):
        # Format index number with leading zeros preserved
        index_number = f"30411000{str(i).zfill(3)}"
        name = f"Student {i}"
        sex = random.choice(["Male", "Female"])
        class_name = random.choice(classes)
        programme = random.choice(programmes)
        
        # Assign subjects based on programme
        student_c_subjects = c_subjects.copy()
        
        # Select 4 elective subjects
        student_e_subjects = random.sample(e_subjects, 4)
        
        data.append([
            class_name, programme, index_number, name, sex,
            student_c_subjects[0], student_c_subjects[1], student_c_subjects[2], student_c_subjects[3],
            student_e_subjects[0], student_e_subjects[1], student_e_subjects[2], student_e_subjects[3]
        ])
    
    columns = [
        "Class", "PROGRAMMES", "INDEX NUMBER", "NAME", "Sex",
        "C-SUBJECT 1", "C-SUBJECT 2", "C-SUBJECT 3", "C-SUBJECT 4",
        "E-SUBJECT 1", "E-SUBJECT 2", "E-SUBJECT 3", "E-SUBJECT 4"
    ]
    
    return pd.DataFrame(data, columns=columns)

def generate_student_scores(df, min_score, max_score):
    """Generate scores for all students and their subjects"""
    # Dictionary to store scores: {index_number: {subject: [year1, year2, year3]}}
    student_scores = {}
    
    for idx, student in df.iterrows():
        index_number = student['INDEX NUMBER']
        student_scores[index_number] = {}
        
        # Get subject columns and their actual values for this student
        subject_columns = df.columns[5:]  # Skip Class, PROGRAMMES, INDEX NUMBER, NAME, Sex
        
        # Generate scores for each subject
        for col in subject_columns:
            if pd.notna(student[col]) and student[col] != "":
                subject_name = student[col]
                # Generate scores for 3 years
                student_scores[index_number][subject_name] = [
                    random.randint(min_score, max_score),
                    random.randint(min_score, max_score),
                    random.randint(min_score, max_score)
                ]
    
    return student_scores

def create_student_sheet(ws, students_df, student_scores, start_row=2):
    """Create a sheet with student data and scores"""
    current_row = start_row
    
    # Process each student
    for idx, student in students_df.iterrows():
        index_number = student['INDEX NUMBER']
        
        # Add student name in the first row of their section
        ws[f'A{current_row}'] = student['NAME']
        
        # Add class, programme, and index number in the next rows
        ws[f'A{current_row+1}'] = f"Class: {student['Class']}"
        ws[f'A{current_row+2}'] = f"Programme: {student['PROGRAMMES']}"
        ws[f'A{current_row+3}'] = f"Index No: {index_number}"
        
        # Get subject columns and their actual values for this student
        subject_columns = students_df.columns[5:]  # Skip Class, PROGRAMMES, INDEX NUMBER, NAME, Sex
        
        # Create a list of actual subjects this student is taking
        student_subjects = []
        for col in subject_columns:
            if pd.notna(student[col]) and student[col] != "":
                student_subjects.append((col, student[col]))  # Store both column name and actual subject
        
        # Add subjects with serial numbers
        for i, (col, subject_name) in enumerate(student_subjects, 1):
            row_offset = i - 1  # Offset for the current subject row
            
            # Calculate the actual row for this subject
            subject_row = current_row + row_offset
            
            # Set S/N
            ws[f'B{subject_row}'] = i
            
            # Set subject name (use the actual subject name from the data)
            ws[f'C{subject_row}'] = subject_name
            
            # Use the pre-generated scores for this student and subject
            if index_number in student_scores and subject_name in student_scores[index_number]:
                scores = student_scores[index_number][subject_name]
                ws[f'D{subject_row}'] = scores[0]  # Year 1
                ws[f'E{subject_row}'] = scores[1]  # Year 2
                ws[f'F{subject_row}'] = scores[2]  # Year 3
            
            # If we've gone beyond the student details rows, leave those cells empty
            if row_offset >= 4:
                ws[f'A{subject_row}'] = ""
        
        # Move to next student (current row + number of subjects + blank row)
        current_row += max(len(student_subjects), 4) + 1
    
    return current_row

def setup_sheet_headers(ws):
    """Set up headers for a student sheet"""
    # Add headers
    ws['A1'] = "Student Details"
    ws['B1'] = "S/N"
    ws['C1'] = "Subjects"
    ws['D1'] = "Year 1"
    ws['E1'] = "Year 2"
    ws['F1'] = "Year 3"
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def process_uploaded_data(uploaded_file, min_score, max_score):
    """Process uploaded Excel file and add random scores"""
    # Read the uploaded Excel file with dtype=str to preserve leading zeros
    df = pd.read_excel(uploaded_file, dtype=str)
    
    # Generate scores for all students once
    student_scores = generate_student_scores(df, min_score, max_score)
    
    # Create a workbook
    wb = openpyxl.Workbook()
    
    # Create the All Students sheet
    all_students_ws = wb.active
    all_students_ws.title = "All Students"
    setup_sheet_headers(all_students_ws)
    create_student_sheet(all_students_ws, df, student_scores)
    
    # Create sheets for each programme
    programmes = df['PROGRAMMES'].unique()
    
    for programme in programmes:
        # Filter students by programme
        programme_df = df[df['PROGRAMMES'] == programme]
        
        if len(programme_df) > 0:
            # Create a new sheet for this programme
            programme_ws = wb.create_sheet(title=programme)
            setup_sheet_headers(programme_ws)
            create_student_sheet(programme_ws, programme_df, student_scores)
    
    # Add student data to the same workbook
    student_ws = wb.create_sheet(title="All Student Data")
    
    # Write headers
    for col_idx, column_name in enumerate(df.columns, 1):
        student_ws.cell(row=1, column=col_idx).value = column_name
        student_ws.cell(row=1, column=col_idx).font = Font(bold=True)
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            student_ws.cell(row=row_idx, column=col_idx).value = value
    
    return wb, df

def create_subject_scores_sheet(min_score=0, max_score=100, num_students=20):
    """Create a sheet with subjects and random scores for three years"""
    # Generate student data first
    df = generate_student_data(num_students)
    
    # Generate scores for all students once
    student_scores = generate_student_scores(df, min_score, max_score)
    
    # Create a workbook
    wb = openpyxl.Workbook()
    
    # Create the All Students sheet
    all_students_ws = wb.active
    all_students_ws.title = "All Students"
    setup_sheet_headers(all_students_ws)
    create_student_sheet(all_students_ws, df, student_scores)
    
    # Create sheets for each programme
    programmes = df['PROGRAMMES'].unique()
    
    for programme in programmes:
        # Filter students by programme
        programme_df = df[df['PROGRAMMES'] == programme]
        
        if len(programme_df) > 0:
            # Create a new sheet for this programme
            programme_ws = wb.create_sheet(title=programme)
            setup_sheet_headers(programme_ws)
            create_student_sheet(programme_ws, programme_df, student_scores)
    
    # Add student data to the same workbook
    student_ws = wb.create_sheet(title="All Student Data")
    
    # Write headers
    for col_idx, column_name in enumerate(df.columns, 1):
        student_ws.cell(row=1, column=col_idx).value = column_name
        student_ws.cell(row=1, column=col_idx).font = Font(bold=True)
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            student_ws.cell(row=row_idx, column=col_idx).value = value
    
    return wb, df

def get_download_link(wb, filename):
    """Generate a download link for the Excel file"""
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)
    b64 = base64.b64encode(virtual_file.read()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'

def main():
    st.set_page_config(page_title="Student Score Generator", layout="wide")
    run_app()

def run_app():
    st.title("Student Score Generator")
    st.write("Generate random scores for students across different years")
    
    # Add template download section
    st.subheader("Need a template?")
    if st.button("Generate Template File"):
        template_wb = create_template_file()
        template_file = io.BytesIO()
        template_wb.save(template_file)
        template_file.seek(0)
        
        b64_template = base64.b64encode(template_file.read()).decode()
        template_download_link = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_template}" download="student_score_generator_template.xlsx">Download Template Excel File</a>'
        st.markdown(template_download_link, unsafe_allow_html=True)
        st.info("Template file generated! Click the link above to download.")
        st.markdown("---")
    
    # Create tabs for different options
    tab1, tab2 = st.tabs(["Upload Data", "Generate New Data"])
    
    with tab1:
        st.header("Upload your Excel file with student data")
        st.write("Your Excel file should include columns: Class, PROGRAMMES, INDEX NUMBER, NAME, Sex, and subject columns")
        uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
        
        if uploaded_file is not None:
            st.success("File uploaded successfully!")
            
            col1, col2 = st.columns(2)
            with col1:
                min_score = st.number_input("Minimum score", min_value=0, max_value=100, value=0, key="min1")
            with col2:
                max_score = st.number_input("Maximum score", min_value=0, max_value=100, value=100, key="max1")
            
            if st.button("Process Uploaded Data"):
                with st.spinner("Processing data..."):
                    try:
                        wb, df = process_uploaded_data(uploaded_file, min_score, max_score)
                        
                        # Get unique programmes for display
                        programmes = df['PROGRAMMES'].unique()
                        
                        # Provide download link
                        download_link = get_download_link(wb, "processed_student_scores.xlsx")
                        st.markdown(download_link, unsafe_allow_html=True)
                        
                        # Preview the uploaded data
                        st.subheader("Preview of Uploaded Student Data")
                        st.dataframe(df.head())
                        
                        # Show information about the generated file
                        st.success(f"Excel file generated with all {len(df)} students on a single sheet.")
                        st.info(f"Additional sheets created for each programme: {', '.join(programmes)}")
                        st.info("Each student's data is displayed with their subjects and scores across 3 years.")
                    except Exception as e:
                        st.error(f"Error processing file: {str(e)}")
    
    with tab2:
        st.header("Generate new student data")
        
        col1, col2 = st.columns(2)
        with col1:
            min_score = st.number_input("Minimum score", min_value=0, max_value=100, value=0, key="min2")
        with col2:
            max_score = st.number_input("Maximum score", min_value=0, max_value=100, value=100, key="max2")
        
        num_students = st.slider("Number of students", min_value=1, max_value=20, value=5)
        
        if st.button("Generate Excel File"):
            with st.spinner("Generating Excel file..."):
                wb, df = create_subject_scores_sheet(min_score, max_score, num_students)
                
                # Get unique programmes for display
                programmes = df['PROGRAMMES'].unique()
                
                # Provide download link
                download_link = get_download_link(wb, "student_scores.xlsx")
                st.markdown(download_link, unsafe_allow_html=True)
                
                # Preview student data
                st.subheader("Preview of Student Data")
                st.dataframe(df.head()) 
                
                # Show information about the generated file
                st.success(f"Excel file generated with all {num_students} students on a single sheet.")
                st.info(f"Additional sheets created for each programme: {', '.join(programmes)}")
                st.info("Each student's data is displayed with their subjects and scores across 3 years.")

    # Add footer with link to services
    st.markdown("---")
    st.markdown(
        """
        <div style="text-align: center; padding: 10px;">
            <p>For more educational tools and services, visit <a href="https://techmawu.com" target="_blank">techmawu.com</a></p>
        </div>
        """, 
        unsafe_allow_html=True
    )

def create_template_file():
    """Creates a template Excel file with the required columns and example data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Define headers
    headers = ["Class", "PROGRAMMES", "INDEX NUMBER", "NAME", "Sex", 
               "C-SUBJECT 1", "C-SUBJECT 2", "C-SUBJECT 3", "C-SUBJECT 4",
               "E-SUBJECT 1", "E-SUBJECT 2", "E-SUBJECT 3", "E-SUBJECT 4"]
    
    # Add headers to the first row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add example data
    example_data = [
        ["3A1", "GENERAL ARTS", "3041100001", "John Doe", "Male", 
         "C-MATHS", "ENG", "SOC-STUD", "INT-SCI", 
         "GOV", "ECONS", "LIT-ENG", "FANTE"],
        
        ["3A2", "SCIENCE", "3041100002", "Jane Smith", "Female", 
         "C-MATHS", "ENG", "SOC-STUD", "INT-SCI", 
         "PHYSICS", "CHEMISTRY", "BIOLOGY", "ECONS"],
        
        ["3S1", "BUSINESS", "3041100003", "Alex Johnson", "Male", 
         "C-MATHS", "ENG", "SOC-STUD", "INT-SCI", 
         "ECONS", "GOV", "GEOGRAPHY", "LIT-ENG"],
        
        ["3S2", "HOME ECONOMICS", "3041100004", "Sarah Williams", "Female", 
         "C-MATHS", "ENG", "SOC-STUD", "INT-SCI", 
         "FANTE", "CHEMISTRY", "BIOLOGY", "GEOGRAPHY"],
        
        ["3B1", "GENERAL ARTS", "3041100005", "Michael Brown", "Male", 
         "C-MATHS", "ENG", "SOC-STUD", "INT-SCI", 
         "GOV", "ECONS", "GEOGRAPHY", "LIT-ENG"]
    ]
    
    # Add example data starting from row 2
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add instructions in a new sheet
    ws_instructions = wb.create_sheet(title="Instructions")
    instructions = [
        ["Instructions for filling the template:"],
        [""],
        ["1. Class: The class designation of the student (e.g., 3A1, 3S2)"],
        ["2. PROGRAMMES: The programme the student is enrolled in (e.g., GENERAL ARTS, SCIENCE)"],
        ["3. INDEX NUMBER: Unique identifier for each student (e.g., 3041100001)"],
        ["4. NAME: Full name of the student"],
        ["5. Sex: Gender of the student (Male/Female)"],
        ["6. C-SUBJECT 1-4: Core subjects taken by the student"],
        ["7. E-SUBJECT 1-4: Elective subjects taken by the student"],
        [""],
        ["Notes:"],
        ["- Do not change the column headers"],
        ["- All students should have 4 core subjects and 4 elective subjects"],
        ["- Make sure subject names are consistent (e.g., 'C-MATHS' vs 'MATHS')"],
        ["- You can add as many rows as needed for additional students"],
        ["- Save the file as Excel (.xlsx) format before uploading"]
    ]
    
    for row_idx, row_text in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=row_text[0])
    
    # Format the instructions
    ws_instructions.column_dimensions['A'].width = 80
    for i in range(1, len(instructions) + 1):
        cell = ws_instructions.cell(row=i, column=1)
        if i == 1:  # Title row
            cell.font = Font(bold=True, size=14)
        elif i > 2 and i <= 9:  # Column descriptions
            cell.font = Font(bold=False)
        elif i > 10:  # Notes
            cell.font = Font(italic=True)
    
    # Set column widths in template sheet
    column_widths = {
        'A': 15,  # Class
        'B': 20,  # PROGRAMMES
        'C': 20,  # INDEX NUMBER
        'D': 30,  # NAME
        'E': 15,  # Sex
        'F': 15,  # C-SUBJECT 1
        'G': 15,  # C-SUBJECT 2
        'H': 15,  # C-SUBJECT 3
        'I': 15,  # C-SUBJECT 4
        'J': 15,  # E-SUBJECT 1
        'K': 15,  # E-SUBJECT 2
        'L': 15,  # E-SUBJECT 3
        'M': 15,  # E-SUBJECT 4
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    return wb

if __name__ == "__main__":
    main()
