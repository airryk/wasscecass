import pandas as pd
import openpyxl
import streamlit as st
import io
import base64
import logging

# Set up logging for debugging
logging.basicConfig(level=logging.INFO)

def reorder_subjects(wb):
    """Reorder subjects in the specified order in all sheets"""
    # Define the core subject order
    core_subject_order = ["SOC-STUD", "ENG", "C-MATHS", "INT-SCI"]
    
    # Process each sheet in the workbook
    for sheet_name in wb.sheetnames:
        # Skip the raw data sheet
        if sheet_name == "All Student Data":
            continue
            
        ws = wb[sheet_name]
        st.write(f"Processing sheet: {sheet_name}")
        
        # First, let's analyze the sheet structure
        st.write(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Display the first few rows to understand the structure
        sample_data = []
        for row in range(1, min(10, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(7, ws.max_column + 1)):
                row_data.append(ws.cell(row=row, column=col).value)
            sample_data.append(row_data)
        
        st.write("Sample data from sheet:")
        st.write(pd.DataFrame(sample_data))
        
        # Try a different approach to find students
        students_processed = 0
        row = 2  # Start from row 2 (after headers)
        
        while row <= ws.max_row:
            # Look for patterns that might indicate a student record
            # Check if column C has a subject and column D, E, F have scores
            subject = ws.cell(row=row, column=3).value
            score1 = ws.cell(row=row, column=4).value
            score2 = ws.cell(row=row, column=5).value
            score3 = ws.cell(row=row, column=6).value
            
            # If we have a subject and at least one score, collect all subjects for this student
            if subject and (score1 or score2 or score3):
                # Find the student name (might be in the current row or above)
                student_name = None
                for r in range(row, max(1, row-5), -1):
                    name_candidate = ws.cell(row=r, column=1).value
                    if name_candidate and not ws.cell(row=r, column=2).value:
                        student_name = name_candidate
                        break
                
                if not student_name:
                    student_name = f"Unknown Student at row {row}"
                
                st.write(f"Found student: {student_name} at row {row}")
                students_processed += 1
                
                # Collect all subjects for this student
                subject_rows = []
                subject_data = []
                current_subject_row = row
                
                # Keep collecting subjects until we hit a row without a subject
                while (current_subject_row <= ws.max_row and 
                       ws.cell(row=current_subject_row, column=3).value):
                    
                    sn = ws.cell(row=current_subject_row, column=2).value
                    subject_name = ws.cell(row=current_subject_row, column=3).value
                    year1 = ws.cell(row=current_subject_row, column=4).value
                    year2 = ws.cell(row=current_subject_row, column=5).value
                    year3 = ws.cell(row=current_subject_row, column=6).value
                    
                    subject_rows.append(current_subject_row)
                    subject_data.append({
                        'row': current_subject_row,
                        'subject': subject_name,
                        'year1': year1,
                        'year2': year2,
                        'year3': year3
                    })
                    
                    current_subject_row += 1
                
                if subject_data:
                    # Sort subjects: first core subjects in specified order, then others alphabetically
                    def subject_sort_key(item):
                        subject_name = item['subject']
                        if subject_name in core_subject_order:
                            return (0, core_subject_order.index(subject_name))
                        return (1, subject_name)
                    
                    sorted_subjects = sorted(subject_data, key=subject_sort_key)
                    
                    # Update the sheet with the new order
                    for i, subject_info in enumerate(sorted_subjects):
                        target_row = subject_rows[i]
                        ws.cell(row=target_row, column=3).value = subject_info['subject']
                        ws.cell(row=target_row, column=4).value = subject_info['year1']
                        ws.cell(row=target_row, column=5).value = subject_info['year2']
                        ws.cell(row=target_row, column=6).value = subject_info['year3']
                    
                    st.write(f"Reordered {len(subject_data)} subjects for {student_name}")
                
                # Move to the row after the last subject
                row = current_subject_row
            else:
                row += 1
        
        st.write(f"Processed {students_processed} students in sheet {sheet_name}")
    
    return wb

def get_download_link(wb, filename):
    """Generate a download link for the Excel file"""
    virtual_file = io.BytesIO()
    wb.save(virtual_file)
    virtual_file.seek(0)
    b64 = base64.b64encode(virtual_file.read()).decode()
    return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'

def main():
    st.set_page_config(page_title="Subject Reordering Tool", layout="wide")
    
    st.title("Subject Reordering Tool")
    st.write("Reorder subjects in the specified order: SOC-STUD, ENG, C-MATHS, INT-SCI, followed by other subjects")
    
    st.header("Upload your generated Excel file")
    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
    
    if uploaded_file is not None:
        st.success("File uploaded successfully!")
        
        if st.button("Reorder Subjects"):
            with st.spinner("Reordering subjects..."):
                try:
                    # Load the workbook
                    file_bytes = uploaded_file.read()
                    virtual_file = io.BytesIO(file_bytes)
                    wb = openpyxl.load_workbook(virtual_file)
                    
                    # Reorder subjects
                    wb = reorder_subjects(wb)
                    
                    # Provide download link
                    download_link = get_download_link(wb, "reordered_student_scores.xlsx")
                    st.markdown(download_link, unsafe_allow_html=True)
                    
                    st.success("Subjects have been reordered successfully!")
                    st.info("Core subjects are now in the order: SOC-STUD, ENG, C-MATHS, INT-SCI, followed by other subjects.")
                    
                except Exception as e:
                    st.error(f"Error processing file: {str(e)}")
                    st.exception(e)  # This will show the full traceback

if __name__ == "__main__":
    main()
